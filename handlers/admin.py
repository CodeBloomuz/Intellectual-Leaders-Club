import logging
from io import BytesIO
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters import Command
import openpyxl

import database as db
import keyboards as kb
from config import ADMIN_IDS

router = Router()
logger = logging.getLogger(__name__)


# ─── Helpers ─────────────────────────────────────────────────────────────────

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


STATUS_UZ = {
    "pending":   "⏳ Kutilmoqda",
    "invited":   "📅 Suhbatga taklif",
    "accepted":  "✅ Qabul qilindi",
    "rejected":  "❌ Rad etildi",
    "probation": "⏳ Sinov muddati",
}


# ─── States ──────────────────────────────────────────────────────────────────

class InterviewForm(StatesGroup):
    date     = State()
    time     = State()
    location = State()
    confirm  = State()


class BroadcastForm(StatesGroup):
    text    = State()
    confirm = State()


# ─── /admin ──────────────────────────────────────────────────────────────────

@router.message(Command("admin"))
async def cmd_admin(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.clear()
    stats = await db.get_stats()
    await message.answer(
        f"🛡️ <b>Admin panel</b>\n\n"
        f"📝 Jami arizalar: <b>{stats['total']}</b>\n"
        f"⏳ Kutilmoqda: <b>{stats['pending']}</b>\n"
        f"📅 Suhbatga taklif: <b>{stats['invited']}</b>\n"
        f"✅ Qabul qilingan: <b>{stats['accepted']}</b>\n"
        f"❌ Rad etilgan: <b>{stats['rejected']}</b>\n"
        f"⏳ Sinov muddati: <b>{stats['probation']}</b>",
        parse_mode="HTML",
        reply_markup=kb.admin_menu_kb()
    )


# ─── Arizalar ro'yxati ────────────────────────────────────────────────────────

@router.message(F.text == "📋 Arizalar ro'yxati")
async def applications_list(message: Message):
    if not is_admin(message.from_user.id):
        return

    users = await db.get_all_users(status="pending")
    if not users:
        await message.answer("⏳ Hozircha kutilayotgan arizalar yo'q.")
        return

    text = f"📋 <b>Kutilayotgan arizalar ({len(users)} ta):</b>\n\n"
    for i, u in enumerate(users, 1):
        text += (
            f"{i}. <b>{u['full_name']}</b>\n"
            f"   🏫 {u['fakultet']}\n"
            f"   👥 {u['guruh']} | 📞 {u['phone']}\n"
            f"   🆔 <code>{u['telegram_id']}</code>\n\n"
        )

    # Telegram max 4096 belgi
    if len(text) > 4000:
        text = text[:4000] + "\n..."

    await message.answer(text, parse_mode="HTML")


# ─── Statistika ───────────────────────────────────────────────────────────────

@router.message(F.text == "📊 Statistika")
async def statistics(message: Message):
    if not is_admin(message.from_user.id):
        return
    stats = await db.get_stats()
    await message.answer(
        f"📊 <b>Statistika</b>\n\n"
        f"📝 Jami: <b>{stats['total']}</b>\n"
        f"⏳ Kutilmoqda: <b>{stats['pending']}</b>\n"
        f"📅 Suhbatga taklif: <b>{stats['invited']}</b>\n"
        f"✅ Qabul: <b>{stats['accepted']}</b>\n"
        f"❌ Rad: <b>{stats['rejected']}</b>\n"
        f"⏳ Sinov muddati: <b>{stats['probation']}</b>",
        parse_mode="HTML"
    )


# ─── Suhbat belgilash ─────────────────────────────────────────────────────────

@router.message(F.text == "📅 Suhbat belgilash")
async def schedule_interview(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    pending = await db.get_all_users(status="pending")
    if not pending:
        await message.answer("⏳ Kutilayotgan nomzodlar yo'q.")
        return

    await state.set_state(InterviewForm.date)
    await message.answer(
        f"📅 Suhbat sanasini kiriting:\n<i>Misol: 03.05.2026</i>",
        parse_mode="HTML",
        reply_markup=kb.cancel_kb()
    )


@router.message(InterviewForm.date)
async def interview_date(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=kb.admin_menu_kb())
        return
    await state.update_data(date=message.text.strip())
    await state.set_state(InterviewForm.time)
    await message.answer(
        "🕙 Suhbat vaqtini kiriting:\n<i>Misol: 10:00</i>",
        parse_mode="HTML"
    )


@router.message(InterviewForm.time)
async def interview_time(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=kb.admin_menu_kb())
        return
    await state.update_data(time=message.text.strip())
    await state.set_state(InterviewForm.location)
    await message.answer(
        "📍 Suhbat joyini kiriting:\n<i>Misol: 3-qavat, 322-xona</i>",
        parse_mode="HTML"
    )


@router.message(InterviewForm.location)
async def interview_location(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=kb.admin_menu_kb())
        return
    await state.update_data(location=message.text.strip())
    data = await state.get_data()
    pending = await db.get_all_users(status="pending")

    await state.set_state(InterviewForm.confirm)
    await message.answer(
        f"📋 <b>Suhbat ma'lumotlari:</b>\n\n"
        f"📅 Sana: <b>{data['date']}</b>\n"
        f"🕙 Soat: <b>{data['time']}</b>\n"
        f"📍 Joy: <b>{data['location']}</b>\n\n"
        f"👥 Xabar yuboriladi: <b>{len(pending)} ta</b> nomzodga\n\n"
        f"Tasdiqlaysizmi?",
        parse_mode="HTML",
        reply_markup=kb.interview_confirm_kb()
    )


@router.callback_query(F.data == "send_interview")
async def send_interview(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    data = await state.get_data()
    pending = await db.get_all_users(status="pending")

    sent = 0
    for u in pending:
        first_name = u["full_name"].split()[0]
        try:
            await callback.bot.send_message(
                u["telegram_id"],
                f"Assalomu alaykum, <b>{first_name}</b>! 👋\n\n"
                f"Siz <b>\"Intellectual Leaders Club\"</b>ga topshirgan "
                f"arizangiz ko'rib chiqildi.\n\n"
                f"📌 <b>Suhbatga taklif etilasiz!</b>\n\n"
                f"📅 Sana: <b>{data['date']}</b>\n"
                f"🕙 Soat: <b>{data['time']}</b>\n"
                f"📍 Joy: Termiz iqtisodiyot va servis universiteti, "
                f"<b>{data['location']}</b>\n\n"
                f"Iltimos, o'z vaqtida tashrif buyuring.",
                parse_mode="HTML",
                reply_markup=kb.interview_response_kb()
            )
            await db.update_status(u["telegram_id"], "invited")
            sent += 1
        except Exception as e:
            logger.warning(f"Yuborishda xato {u['telegram_id']}: {e}")

    await db.save_interview(data["date"], data["time"], data["location"])
    await callback.message.answer(
        f"✅ Suhbat xabari <b>{sent}</b> ta nomzodga yuborildi!",
        parse_mode="HTML",
        reply_markup=kb.admin_menu_kb()
    )
    await state.clear()
    await callback.answer()


@router.callback_query(F.data == "cancel_interview")
async def cancel_interview(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.answer("Bekor qilindi.", reply_markup=kb.admin_menu_kb())
    await callback.answer()


# ─── Qabul / Sinov / Rad ─────────────────────────────────────────────────────

@router.callback_query(F.data.startswith("accept_"))
async def accept_user(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    tid = int(callback.data.split("_")[1])
    await db.update_status(tid, "accepted")
    try:
        await callback.bot.send_message(
            tid,
            "🎉 <b>Tabriklaymiz!</b>\n\n"
            "Siz <b>\"Intellectual Leaders Club\"</b>ga a'zo bo'ldingiz!\n"
            "Tez orada klub faoliyati haqida xabar beramiz. 🚀",
            parse_mode="HTML"
        )
    except Exception:
        pass
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer("✅ Nomzod qabul qilindi!")
    await callback.answer()


@router.callback_query(F.data.startswith("probation_"))
async def probation_user(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    tid = int(callback.data.split("_")[1])
    await db.update_status(tid, "probation")
    try:
        await callback.bot.send_message(
            tid,
            "📋 <b>Sinov muddati</b>\n\n"
            "Siz <b>\"Intellectual Leaders Club\"</b>ga sinov muddatiga "
            "qabul qilindingiz!\n"
            "Sinov muddati: <b>1 oy</b>\n\n"
            "Bu davr mobaynida faolligingizni ko'rsating! 💪",
            parse_mode="HTML"
        )
    except Exception:
        pass
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer("⏳ Nomzod sinov muddatiga qo'yildi!")
    await callback.answer()


@router.callback_query(F.data.startswith("reject_"))
async def reject_user(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    tid = int(callback.data.split("_")[1])
    await db.update_status(tid, "rejected")
    try:
        await callback.bot.send_message(
            tid,
            "📋 Arizangiz ko'rib chiqildi.\n\n"
            "Afsuski, bu safar qabul qilinmadi.\n"
            "Kelajakda yana harakat qilishingiz mumkin! 💪"
        )
    except Exception:
        pass
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer("❌ Nomzod rad etildi.")
    await callback.answer()


# ─── Excel eksport ────────────────────────────────────────────────────────────

@router.message(F.text == "📤 Excel eksport")
async def excel_export(message: Message):
    if not is_admin(message.from_user.id):
        return

    users = await db.get_all_users()
    if not users:
        await message.answer("Hozircha arizalar yo'q.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arizalar"

    headers = ["#", "To'liq ism", "Telegram ID", "Fakultet", "Yo'nalish",
               "Guruh", "Telefon", "Qiziqish yo'nalishi", "Motivatsiya", "Holat", "Sana"]
    ws.append(headers)

    # Header formati
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF")

    for i, u in enumerate(users, 1):
        ws.append([
            i,
            u["full_name"],
            u["telegram_id"],
            u["fakultet"],
            u["yonalish"],
            u["guruh"],
            u["phone"],
            u["interest"],
            u["motivation"],
            STATUS_UZ.get(u["status"], u["status"]),
            str(u["created_at"])
        ])

    # Ustun kengligi
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    file = BufferedInputFile(buf.read(), filename="ILC_Arizalar.xlsx")
    await message.answer_document(
        file,
        caption=f"📊 Jami {len(users)} ta ariza"
    )


# ─── Broadcast ────────────────────────────────────────────────────────────────

@router.message(F.text == "📢 Broadcast")
async def broadcast_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    accepted = await db.get_all_users(status="accepted")
    await state.set_state(BroadcastForm.text)
    await message.answer(
        f"📢 Barcha qabul qilingan a'zolarga (<b>{len(accepted)} kishi</b>) "
        f"yuboriladigan xabarni yozing:",
        parse_mode="HTML",
        reply_markup=kb.cancel_kb()
    )


@router.message(BroadcastForm.text)
async def broadcast_text(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=kb.admin_menu_kb())
        return
    await state.update_data(broadcast_msg=message.text)
    await state.set_state(BroadcastForm.confirm)
    await message.answer(
        f"📢 <b>Xabar:</b>\n\n{message.text}\n\n"
        f"Yuborilsinmi?",
        parse_mode="HTML",
        reply_markup=kb.broadcast_confirm_kb()
    )


@router.callback_query(F.data == "send_broadcast")
async def send_broadcast(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    data = await state.get_data()
    msg = data["broadcast_msg"]

    accepted = await db.get_all_users(status="accepted")
    sent = 0
    for u in accepted:
        try:
            await callback.bot.send_message(u["telegram_id"], msg)
            sent += 1
        except Exception:
            pass

    await callback.message.answer(
        f"✅ Xabar <b>{sent}</b> ta a'zoga yuborildi!",
        parse_mode="HTML",
        reply_markup=kb.admin_menu_kb()
    )
    await state.clear()
    await callback.answer()


@router.callback_query(F.data == "cancel_broadcast")
async def cancel_broadcast(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.answer("Bekor qilindi.", reply_markup=kb.admin_menu_kb())
    await callback.answer()
